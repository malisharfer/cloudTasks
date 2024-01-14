from openpyxl import load_workbook
from azure.storage.blob import BlobServiceClient
import base64
import io
import os
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import config.config_variables
import project.get_connection_string


def download_blob(container_name, blob_name):
    secret = config.config_variables.email_secret
    connection_string = project.get_connection_string.get_connection_string_from_keyvault(secret)
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    container_client = blob_service_client.get_container_client(container_name)
    blob_client = container_client.get_blob_client(blob_name)
    blob_data = blob_client.download_blob().readall()
    return blob_data


def write_to_excel(subscription_obj):
    secret = config.config_variables.email_secret
    connection_string = project.get_connection_string.get_connection_string_from_keyvault(secret)
    container_name = 'excel'
    blob_name = 'file_subscription.xlsx'
    excel_sheet = download_blob(container_name, blob_name)
    encoded_data = base64.b64encode(excel_sheet).decode('utf-8')
    bytes_data = base64.b64decode(encoded_data)
    workbook = load_workbook(io.BytesIO(bytes_data))
    sheet = workbook.active
    global last_cell_written
    last_cell_written = sheet.max_row
    file_stream = io.BytesIO()
    column_subscription_name = "A{}".format(last_cell_written + 1)
    sheet[column_subscription_name] = subscription_obj['display_name']
    column_subscription_id = "B{}".format(last_cell_written + 1)
    sheet[column_subscription_id] = subscription_obj['subscription_id']
    column_body = "C{}".format(last_cell_written + 1)
    sheet[column_body] = subscription_obj['body']
    workbook.save(file_stream)
    file_stream.seek(0)
    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name)
    blob_client.upload_blob(file_stream.getvalue(), overwrite = True)
